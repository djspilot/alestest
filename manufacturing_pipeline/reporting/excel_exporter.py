"""Excel export in SpaceClaim-compatible column format.

Exports pipeline analysis results to .xlsx with the same 26-column layout
as SpaceClaim/AutoPOL, enabling direct side-by-side comparison.

Supports optional reference data (SpaceClaim Excel or XML) for a
comparison sheet with color-coded deviations.
"""

import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# SpaceClaim column definition (exact match with their Excel export)
# ---------------------------------------------------------------------------

SPACECLAIM_COLUMNS = [
    "Artikel-import",  # 1  - Part name
    "Aantal",          # 2  - Quantity
    "Soort",           # 3  - Kind (plaat/koker/anders/Samenstel)
    "Materiaal",       # 4  - Material code
    "Artikel code",    # 5  - Article code (same as part name)
    "Dikte",           # 6  - Thickness (mm)
    "Materiaal",       # 7  - Material (duplicate column in SpaceClaim)
    "Vorm",            # 8  - Shape (plaat/koker/buis)
    "Opmerking",       # 9  - Notes / warnings
    "Foutmelding",     # 10 - Error messages
    "Oppervlakte",     # 11 - Surface area (mm²)
    "Snijlengte",      # 12 - Total cut length (mm)
    "Lengte",          # 13 - Length (flat or bbox, mm)
    "Breedte",         # 14 - Width (flat or bbox, mm)
    "Hoogte/Dia",      # 15 - Height or diameter (mm)
    "Snijgaten",       # 16 - Laser-cut holes count
    "Aantalzet",       # 17 - Bend count (for ERP)
    "tegenzet",        # 18 - Counter-bends
    "Boorgaten",       # 19 - Drilled holes count
    "DiaBoorgat",      # 20 - Drill hole diameter
    "OppervNetto",     # 21 - Net surface area (mm²)
    "Gewicht",         # 22 - Weight (kg)
    "Nabewerking",     # 23 - Post-processing
    "offertenr",       # 24 - Quote number
    "aantal bewerk",   # 25 - Processing count
    "PDF",             # 26 - PDF available
]

# Comparison fields (numeric columns worth comparing)
_COMPARE_FIELDS = [
    ("Dikte", 5),
    ("Oppervlakte", 10),
    ("Snijlengte", 11),
    ("Lengte", 12),
    ("Breedte", 13),
    ("Hoogte/Dia", 14),
    ("Snijgaten", 15),
    ("Aantalzet", 16),
    ("tegenzet", 17),
    ("Gewicht", 21),
]

# Cell fills for comparison sheet
_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_ORANGE = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FONT_HEADER = Font(bold=True)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_to_excel(
    results: List[Dict[str, Any]],
    output_path: Path,
    reference_path: Optional[Path] = None,
) -> None:
    """Export analysis results to Excel in SpaceClaim column format.

    Args:
        results: List of pipeline result dicts (from process_single_file or API).
        output_path: Destination .xlsx path.
        reference_path: Optional path to SpaceClaim reference (.xlsx or .xml)
                        for side-by-side comparison sheet.
    """
    wb = Workbook()

    # Sheet 1: pipeline results in SpaceClaim format
    ws = wb.active
    ws.title = "pipeline"
    _write_pipeline_sheet(ws, results)

    # Sheet 2 (optional): comparison with reference
    if reference_path and reference_path.exists():
        ref_rows = _load_reference(reference_path)
        if ref_rows:
            ws_cmp = wb.create_sheet(title="vergelijking")
            _write_comparison_sheet(ws_cmp, results, ref_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# Pipeline sheet
# ---------------------------------------------------------------------------

def _write_pipeline_sheet(ws, results: List[Dict[str, Any]]) -> None:
    """Write pipeline results using SpaceClaim column headers."""
    # Header row
    for col_idx, header in enumerate(SPACECLAIM_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER

    # Data rows
    for row_idx, result in enumerate(results, 2):
        row_data = _map_result_to_row(result)
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width columns
    _auto_column_widths(ws)


def _map_result_to_row(result: Dict[str, Any]) -> list:
    """Map a pipeline result dict to SpaceClaim column values.

    Handles both the simple format (from process_single_file / batch)
    and the enriched format (from API / analysis_service).
    """
    file_stem = Path(result.get("file", "unknown")).stem

    # Extract nested dicts (handle both formats)
    dims = result.get("dimensions") or {}
    flat = result.get("flat_dimensions") or {}
    prod = result.get("production") or {}
    aag = result.get("aag_details") or {}

    # Holes: try production dict first, then top-level 'holes' key
    holes = prod.get("holes_total") or result.get("holes", 0)

    # Bends: try production dict first, then top-level 'bends' key
    bends = prod.get("bends_total") or result.get("bends", 0)

    # Counter-bends
    counter_bends = prod.get("bends_down", 0)
    if not counter_bends:
        counter_bends = aag.get("counter_bend_count", 0)

    thickness = result.get("thickness", 0) or 0
    category = result.get("category", "")
    part_type = result.get("part_type", "")

    # Dimensions: prefer flat pattern, fall back to bounding box
    length = flat.get("length") or dims.get("length", 0)
    width = flat.get("width") or dims.get("width", 0)
    height = dims.get("height", 0)

    # Surface area estimate (flat area = length * width for sheet metal)
    if flat.get("length") and flat.get("width"):
        oppervlakte = round(flat["length"] * flat["width"], 2)
    elif length and width:
        oppervlakte = round(length * width, 2)
    else:
        oppervlakte = 0

    # Cut length from AAG
    snijlengte = aag.get("total_cut_length", 0) or 0

    # Net area (oppervlakte minus hole areas - rough estimate)
    opperv_netto = oppervlakte  # TODO: subtract hole areas when available

    # Weight estimate (volume * density)
    gewicht = _estimate_weight(dims, thickness)

    # Error / warning messages
    error = result.get("error", "")
    opmerking = ""
    if not result.get("success", True):
        opmerking = error

    return [
        file_stem,                              # 1  Artikel-import
        1,                                      # 2  Aantal
        _soort_from_category(category),         # 3  Soort
        "",                                     # 4  Materiaal
        file_stem,                              # 5  Artikel code
        thickness if thickness else 0,          # 6  Dikte
        "",                                     # 7  Materiaal (2)
        _vorm_from_part_type(part_type),        # 8  Vorm
        opmerking,                              # 9  Opmerking
        error if not result.get("success") else "",  # 10 Foutmelding
        round(oppervlakte, 2),                  # 11 Oppervlakte
        round(snijlengte, 2),                   # 12 Snijlengte
        round(length, 2),                       # 13 Lengte
        round(width, 2),                        # 14 Breedte
        round(height, 2),                       # 15 Hoogte/Dia
        holes,                                  # 16 Snijgaten
        bends,                                  # 17 Aantalzet
        counter_bends if counter_bends else False,  # 18 tegenzet
        0,                                      # 19 Boorgaten
        0,                                      # 20 DiaBoorgat
        round(opperv_netto, 2),                 # 21 OppervNetto
        round(gewicht, 3),                      # 22 Gewicht
        "",                                     # 23 Nabewerking
        "",                                     # 24 offertenr
        0,                                      # 25 aantal bewerk
        False,                                  # 26 PDF
    ]


# ---------------------------------------------------------------------------
# Comparison sheet
# ---------------------------------------------------------------------------

def _write_comparison_sheet(
    ws,
    results: List[Dict[str, Any]],
    ref_rows: Dict[str, list],
) -> None:
    """Write side-by-side comparison: pipeline vs reference with color coding.

    Columns: Artikel | Veld | Pipeline | SpaceClaim | Verschil | Verschil_%
    Color: green = match, orange = <10% off, red = >10% off
    """
    headers = ["Artikel", "Veld", "Pipeline", "SpaceClaim", "Verschil", "Verschil_%"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = _FONT_HEADER

    row = 2
    for result in results:
        file_stem = Path(result.get("file", "")).stem
        pipeline_row = _map_result_to_row(result)

        # Find matching reference row by part name
        ref_row = ref_rows.get(file_stem)
        if ref_row is None:
            # Try partial match (reference may use different naming)
            for ref_key in ref_rows:
                if file_stem in ref_key or ref_key in file_stem:
                    ref_row = ref_rows[ref_key]
                    break

        if ref_row is None:
            ws.cell(row=row, column=1, value=file_stem)
            ws.cell(row=row, column=2, value="(geen referentie gevonden)")
            row += 1
            continue

        for field_name, col_idx in _COMPARE_FIELDS:
            pl_val = _to_float(pipeline_row[col_idx])
            ref_val = _to_float(ref_row[col_idx] if col_idx < len(ref_row) else 0)

            diff = pl_val - ref_val
            diff_pct = ""
            if ref_val != 0:
                diff_pct = round((diff / ref_val) * 100, 1)

            ws.cell(row=row, column=1, value=file_stem)
            ws.cell(row=row, column=2, value=field_name)
            ws.cell(row=row, column=3, value=pl_val)
            ws.cell(row=row, column=4, value=ref_val)
            ws.cell(row=row, column=5, value=round(diff, 2))
            ws.cell(row=row, column=6, value=diff_pct)

            # Color coding
            fill = _get_comparison_fill(pl_val, ref_val)
            for c in range(3, 7):
                ws.cell(row=row, column=c).fill = fill

            row += 1

        # Blank row between parts
        row += 1

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Reference data loading
# ---------------------------------------------------------------------------

def _load_reference(path: Path) -> Optional[Dict[str, list]]:
    """Load SpaceClaim reference data from Excel or XML.

    Returns dict mapping part_name -> row values (same order as SPACECLAIM_COLUMNS).
    """
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return _load_reference_excel(path)
    elif suffix == ".xml":
        return _load_reference_xml(path)
    return None


def _load_reference_excel(path: Path) -> Dict[str, list]:
    """Load reference from SpaceClaim Excel (.xlsx).

    Looks for a sheet named 'spaceclaim' first, then falls back to first sheet.
    """
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return {}

    # Prefer 'spaceclaim' sheet
    if "spaceclaim" in wb.sheetnames:
        ws = wb["spaceclaim"]
    else:
        ws = wb.active

    rows_dict = {}
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header_row = [str(c or "") for c in row]
            continue
        # Use first column (Artikel-import) as key
        part_name = str(row[0] or "").strip()
        if not part_name:
            continue
        rows_dict[part_name] = list(row)

    wb.close()
    return rows_dict


def _load_reference_xml(path: Path) -> Dict[str, list]:
    """Load reference from SpaceClaim XML export.

    Converts XML CalculationResult elements to the same row format.
    """
    try:
        tree = ET.parse(str(path))
        root = tree.getroot()
    except Exception:
        return {}

    rows_dict = {}
    for calc in root.findall(".//CalculationResult"):
        part_name = calc.findtext("Sheet_PartName", "").strip()
        if not part_name:
            continue

        holes = int(calc.findtext("Sheet_NrHoles", "0") or 0)
        bends = int(calc.findtext("Sheet_NrBends", "0") or 0)
        thickness = float(calc.findtext("Sheet_Thickness", "0") or 0)
        length = float(calc.findtext("Sheet_BoxX", "0") or 0)
        width = float(calc.findtext("Sheet_BoxY", "0") or 0)
        height = float(calc.findtext("Sheet_BoxZ", "0") or 0)
        weight = float(calc.findtext("Sheet_Weight", "0") or 0)
        material = calc.findtext("Sheet_Material", "")
        top_area = float(calc.findtext("Sheet_TopArea", "0") or 0)
        outer_contour = float(calc.findtext("Sheet_OuterContour", "0") or 0)
        total_contour = float(calc.findtext("Sheet_TotalContour", "0") or 0)

        # Build row in SPACECLAIM_COLUMNS order
        row = [
            part_name,       # 1  Artikel-import
            1,               # 2  Aantal
            "",              # 3  Soort
            material,        # 4  Materiaal
            part_name,       # 5  Artikel code
            thickness,       # 6  Dikte
            material,        # 7  Materiaal (2)
            "",              # 8  Vorm
            "",              # 9  Opmerking
            "",              # 10 Foutmelding
            top_area,        # 11 Oppervlakte
            total_contour,   # 12 Snijlengte
            length,          # 13 Lengte
            width,           # 14 Breedte
            height,          # 15 Hoogte/Dia
            holes,           # 16 Snijgaten
            bends,           # 17 Aantalzet
            0,               # 18 tegenzet
            0,               # 19 Boorgaten
            0,               # 20 DiaBoorgat
            0,               # 21 OppervNetto
            weight,          # 22 Gewicht
            "",              # 23 Nabewerking
            "",              # 24 offertenr
            0,               # 25 aantal bewerk
            False,           # 26 PDF
        ]
        rows_dict[part_name] = row

    return rows_dict


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _soort_from_category(category: str) -> str:
    """Map pipeline category to SpaceClaim 'Soort' value."""
    cat = (category or "").upper()
    if "PLAAT" in cat or "SHEET" in cat:
        return "plaat"
    if "KOKER" in cat or "PROFIEL" in cat or "TUBE" in cat:
        return "koker"
    if "DRAAI" in cat or "TURNED" in cat:
        return "anders"
    if "ASSEMBLY" in cat or "SAMENSTEL" in cat:
        return "Samenstel"
    return "anders"


def _vorm_from_part_type(part_type: str) -> str:
    """Map pipeline part_type to SpaceClaim 'Vorm' value."""
    pt = (part_type or "").upper()
    if pt in ("PLAAT", "VLAK", "FLAT"):
        return "plaat"
    if pt in ("KOKER", "BUIS"):
        return "koker"
    if pt in ("U_PROFIEL", "HOEKPROFIEL", "L_PROFIEL"):
        return "profiel"
    if "COMPLEX" in pt:
        return "plaat"
    return ""


def _estimate_weight(dims: dict, thickness: float) -> float:
    """Rough weight estimate in kg. Steel density = 7.85 g/cm³."""
    length = dims.get("length", 0) or 0
    width = dims.get("width", 0) or 0

    if thickness > 0 and length > 0 and width > 0:
        # Sheet metal: L × W × T × density
        volume_mm3 = length * width * thickness
        return (volume_mm3 / 1000) * 7.85 / 1000  # g/cm³ → kg
    return 0


def _to_float(val) -> float:
    """Convert value to float, returning 0 on failure."""
    if val is None or val == "" or val is False:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _get_comparison_fill(pipeline_val: float, ref_val: float):
    """Return cell fill color based on deviation between values."""
    if ref_val == 0 and pipeline_val == 0:
        return _FILL_GREEN
    if ref_val == 0:
        return _FILL_RED

    diff_pct = abs((pipeline_val - ref_val) / ref_val) * 100

    if diff_pct <= 1:
        return _FILL_GREEN
    elif diff_pct <= 10:
        return _FILL_ORANGE
    else:
        return _FILL_RED


def _auto_column_widths(ws, min_width: int = 8, max_width: int = 30) -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len
