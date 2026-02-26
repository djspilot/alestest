#!/usr/bin/env python3
"""
Export classification results to Excel
Generates Excel files with: Partnaam, Aantal, Klasse
"""
import sys
import re
from pathlib import Path
import cadquery as cq
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from manufacturing_pipeline.analysis.assembly_analysis import (
    analyze_assembly_complete,
    parse_step_assembly_structure,
    parse_step_product_names,
)

def export_to_excel(step_file, output_file):
    """Export classification results to Excel."""
    
    print(f"\nProcessing: {step_file.name}")
    print("-" * 60)
    
    # Load and analyze
    doc = cq.importers.importStep(str(step_file))
    result = analyze_assembly_complete(
        doc,
        assembly_name="Manufacturing Assembly",
        material="steel_s235",
        step_file_path=str(step_file)
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Classificatie"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Add headers
    headers = ["Partnaam", "Aantal", "Klasse"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Add data rows
    row = 2
    total_quantity = 0
    part_count = 0

    class_map = {
        'plaat': 'Plaat',
        'profiel': 'Profiel',
        'anders': 'Anders'
    }

    # Try to get part names from STEP file
    # Method 1: Assembly structure (NEXT_ASSEMBLY_USAGE_OCCURRENCE)
    step_parts = parse_step_assembly_structure(str(step_file))
    
    # Method 2: Product definitions (PRODUCT_DEFINITION) - ordered list
    step_product_names = None
    if not step_parts:
        step_product_names = parse_step_product_names(str(step_file))
        
        # Deduplicate product names: remove _1, _2, _3 suffixes and keep unique base names
        if step_product_names:
            unique_names = []
            seen_bases = set()
            for name in step_product_names:
                # Remove instance suffix like "_1", "_2", "_3" to get base name
                base_name = re.sub(r'_\d+$', '', name)
                if base_name not in seen_bases:
                    unique_names.append(base_name)
                    seen_bases.add(base_name)
            step_product_names = unique_names
    
    # Create a lookup for BOM items by their part_name
    bom_by_name = {item['part_name']: item for item in result['flat_bom']}
    
    # Track which step_parts we've already used
    used_step_parts = set()
    
    # Base name for generated part names
    base_name = step_file.stem
    generated_idx = 1
    product_name_idx = 0
    
    # Process each BOM item
    for bom_item in result['flat_bom']:
        part_count += 1
        quantity = bom_item['quantity']
        total_quantity += quantity
        
        part_class = bom_item.get('part_class', '')
        if bom_item.get('is_fastener'):
            part_class = 'anders'
        class_name = class_map.get(part_class, 'Anders')
        
        # Determine part name:
        # 1. If BOM item has a name in step_parts (assembly structure), use it
        # 2. Else if we have product names from PRODUCT_DEFINITION, use next one
        # 3. Otherwise, generate a name with -p1, -p2, etc.
        bom_part_name = bom_item.get('part_name', '')
        part_name = None
        
        if step_parts and bom_part_name in step_parts and bom_part_name not in used_step_parts:
            # Use the STEP assembly structure name
            part_name = bom_part_name
            used_step_parts.add(bom_part_name)
        elif step_product_names and product_name_idx < len(step_product_names):
            # Use the next PRODUCT_DEFINITION name (deduplicated)
            part_name = step_product_names[product_name_idx]
            product_name_idx += 1
        
        if not part_name:
            # Generate a new name
            part_name = f"{base_name}-p{generated_idx}"
            generated_idx += 1
        
        # Override classification for DIN/EN/ISO standard profiles (purchased items)
        if part_name:
            name_upper = part_name.upper()
            is_standard = any(std in name_upper for std in ['DIN ', 'DIN-', 'EN ', 'EN-', 'ISO ', 'ISO-'])
            if is_standard:
                # Standard catalog items are classified as "Anders" (not manufactured)
                part_class = 'anders'
                class_name = 'Anders'
        
        # Add row
        ws.cell(row=row, column=1).value = part_name
        ws.cell(row=row, column=2).value = quantity
        ws.cell(row=row, column=3).value = class_name

        # Apply borders and alignment
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col == 2:
                cell.alignment = center_align

        row += 1
    
    # Add totals row
    total_row = row + 1
    ws.cell(row=total_row, column=1).value = f"TOTAAL ({part_count} onderdelen)"
    ws.cell(row=total_row, column=2).value = total_quantity
    ws.cell(row=total_row, column=3).value = ""
    
    # Style totals row
    for col in range(1, 4):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border
        if col == 2:
            cell.alignment = center_align
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    
    # Save
    wb.save(output_file)
    print(f"[OK] Excel saved: {output_file}")
    print(f"  - {part_count} onderdelen")
    print(f"  - {total_quantity} totale items")
    
    return True

def main():
    """Process both test files."""
    workspace = Path(__file__).parent.parent
    test_files = [
        workspace / "stepfiles" / "10040852_1.stp",
        workspace / "stepfiles" / "10040878_1.stp",
        workspace / "stepfiles" / "2006020_A-00.STEP",
        workspace / "stepfiles" / "MD-16-03698_R2.stp",
        workspace / "stepfiles" / "31686-080.stp",
    ]
    
    print("=" * 60)
    print("EXPORT CLASSIFICATIE NAAR EXCEL")
    print("=" * 60)
    
    for step_file in test_files:
        if not step_file.exists():
            print(f"❌ File not found: {step_file}")
            continue
        
        output_file = step_file.parent / f"{step_file.stem}_classificatie.xlsx"
        try:
            export_to_excel(step_file, output_file)
        except Exception as e:
            print(f"[ERROR] {e}")
            import traceback
            traceback.print_exc()
            return False
    
    print("\n" + "=" * 60)
    print("✓ Alle bestanden gegenereerd!")
    print("=" * 60)
    return True

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
