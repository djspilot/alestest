#!/usr/bin/env python3
"""
Final Verification Test - Plate Detection Improvements
Tests all 5 STEP files and verifies classification correctness
"""

import sys
sys.path.insert(0, '.')

from pathlib import Path
from openpyxl import load_workbook

print("="*80)
print("FINAL VERIFICATION TEST - PLATE DETECTION")
print("="*80)

# Expected results based on manual verification
EXPECTED = {
    '10040852_1.stp': {
        'plaat': 2,
        'profiel': 2,
        'anders': 1,
        'total': 5
    },
    '10040878_1.stp': {
        'plaat': 2,
        'profiel': 2,
        'anders': 1,
        'total': 5
    },
    '2006020_A-00.STEP': {
        'plaat': 0,
        'profiel': 0,
        'anders': 2,
        'total': 2
    },
    'MD-16-03698_R2.stp': {
        'plaat': 2,
        'profiel': 1,
        'anders': 2,
        'total': 5
    },
    '31686-080.stp': {
        'plaat': 8,
        'profiel': 0,
        'anders': 10,
        'total': 18
    }
}

# Key verification items
KEY_ITEMS = {
    '31686-080.stp': {
        'DIN 1026 - U 160 - 600': 'Anders',  # Standard profile
        'EN 10210-2 - 88,9 x 4 - 65': 'Anders',  # Standard tube
        '31686-404': 'Plaat',  # Thin plate
        '31686-362': 'Plaat',  # Thin plate
        '31686-380': 'Plaat',  # Thick plate with face detection
    },
    'MD-16-03698_R2.stp': {
        'MD-16-03781_R1': 'Profiel',  # Should be profile, not false positive plate
    }
}

stepfiles_dir = Path('../stepfiles')
all_passed = True

for step_file, expected in EXPECTED.items():
    print(f"\n{'─'*80}")
    print(f"Testing: {step_file}")
    print(f"{'─'*80}")
    
    excel_file = stepfiles_dir / f"{step_file.rsplit('.', 1)[0]}_classificatie.xlsx"
    
    if not excel_file.exists():
        print(f"  ❌ FAIL: Excel file not found")
        all_passed = False
        continue
    
    # Load and analyze
    wb = load_workbook(str(excel_file))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    # Count classifications
    plaat_count = sum(1 for r in rows if r and r[2] == 'Plaat')
    profiel_count = sum(1 for r in rows if r and r[2] == 'Profiel')
    anders_count = sum(1 for r in rows if r and r[2] == 'Anders')
    total = plaat_count + profiel_count + anders_count
    
    # Check totals
    totals_ok = (
        plaat_count == expected['plaat'] and
        profiel_count == expected['profiel'] and
        anders_count == expected['anders'] and
        total == expected['total']
    )
    
    if totals_ok:
        print(f"  ✓ Totals: {total} parts ({plaat_count} Plaat, {profiel_count} Profiel, {anders_count} Anders)")
    else:
        print(f"  ❌ Totals MISMATCH:")
        print(f"     Expected: {expected['total']} total ({expected['plaat']} Plaat, {expected['profiel']} Profiel, {expected['anders']} Anders)")
        print(f"     Got:      {total} total ({plaat_count} Plaat, {profiel_count} Profiel, {anders_count} Anders)")
        all_passed = False
    
    # Check key items if defined for this file
    if step_file in KEY_ITEMS:
        print(f"\n  Key items verification:")
        for part_name, expected_class in KEY_ITEMS[step_file].items():
            found = False
            for r in rows:
                if r and r[0] and part_name in str(r[0]):
                    actual_class = r[2]
                    if actual_class == expected_class:
                        print(f"    ✓ {part_name[:40]}: {actual_class}")
                    else:
                        print(f"    ❌ {part_name[:40]}: Expected {expected_class}, got {actual_class}")
                        all_passed = False
                    found = True
                    break
            if not found:
                print(f"    ❌ {part_name[:40]}: NOT FOUND in Excel")
                all_passed = False

print(f"\n{'='*80}")
if all_passed:
    print("✓ ALL TESTS PASSED")
    print("Classification system working correctly on all 5 STEP files")
else:
    print("❌ SOME TESTS FAILED")
    print("Please review the output above")
print(f"{'='*80}\n")

sys.exit(0 if all_passed else 1)
